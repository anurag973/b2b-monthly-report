"""
B2B Monthly Report.
- When run via GitHub Actions: fetches data, builds Excel, uploads to Drive, sends email
- When imported by Dataiku: run_pipeline() returns a DataFrame for DB write
"""
import json
import os
import smtplib
import time
import traceback
from calendar import monthrange
from datetime import date
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import pandas as pd
import requests

# Only import these when not running in Dataiku
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

# ── Dates (auto-calculated) ────────────────────────────────────────────────
today = date.today()
first = today.replace(day=1)
last_day = monthrange(today.year, today.month)[1]
SNAPSHOT_DATE = first.isoformat()
DUE_START = first.isoformat()
DUE_END = today.replace(day=last_day).isoformat()
MONTH_LABEL = today.strftime("%B_%Y").lower()

# ── Superset ───────────────────────────────────────────────────────────────
SUPERSET_URL = os.environ.get("SUPERSET_URL", "https://superset.bkosh.com")
RISK_DB_ID = 45
KOSH_DB_ID = 1

# ── Paths (only used for GitHub Actions Excel output) ──────────────────────
BASE = Path("output") / f"B2B_{MONTH_LABEL}"
OUT_PATH = BASE / f"b2b_sheet_{SNAPSHOT_DATE}_cleaned.xlsx"

# ── Batch sizes ────────────────────────────────────────────────────────────
PAGE_SIZE = 10000
INTEREST_BATCH = 3000

# ── Email ──────────────────────────────────────────────────────────────────
GMAIL_ADDRESS = os.environ.get("GMAIL_ADDRESS")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD")
REPORT_RECIPIENT = os.environ.get("REPORT_RECIPIENT")

# ── Google Drive ───────────────────────────────────────────────────────────
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID")
GDRIVE_SERVICE_ACCOUNT_JSON = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")


RISK_COLUMNS = [
    "id", "loan_id", "loanshare_id", "amount", "disbursal_date", "name", "mobile",
    "pan_number", "state", "aadhaar_number", "pincode", "address", "partner_loan_id",
    "last_payment_date", "gender", "dob", "age", "employment_type", "annual_salary",
    "topup", "cibil_score", "foir_score", "total_collection", "dpd", "status",
    "tenure", "lender", "purpose", "channel", "user_status", "overdue_amount",
    "principal_outstanding", "tags_location", "relationship_manager", "greypool",
    "emi", "date_created", "last_updated", "writeoff_action", "writeoff_status",
    "fldg", "last_success_payment_date", "overdue_principal",
]

HEADERS = [
    "id", "loan_id", "loanshare_id", "amount", "disbursal_date", "db_month", "name",
    "mobile", "pan_number", "state", "aadhaar_number", "pincode", "address",
    "partner_loan_id", "last_payment_date", "gender", "dob", "age", "employment_type",
    "annual_salary", "topup", "cibil_score", "foir_score", "total_collection", "dpd",
    "status", "tenure", "lender", "purpose", "channel", "user_status", "overdue_amount",
    "principal_outstanding", "tags_location", "relationship_manager", "greypool", "emi",
    "date_created", "last_updated", "writeoff_action", "writeoff_status", "fldg",
    "last_success_payment_date", "overdue_principal", "accrued_interest", "Final POS",
]

REM_HEADERS = [
    "loanshare_id", "loan_id", "amount", "disbursal_date", "name", "mobile",
    "pan_number", "lender", "user_status", "writeoff_status", "writeoff_action",
    "principal_outstanding", "dpd_june", "overdue_amount", "overdue_principal",
    "last_payment_date",
]

FULL_ZERO_LENDERS = {
    "arthmate", "anand_property", "arvog", "unnayan_bharat", "kudos",
    "capital trade link", "kalandri", "koshadmin", "payme_india_pvt_ltd",
    "testlender", "aditsh",
}
PARTIAL_LENDERS_180 = {
    "narendra_finance", "grow_money", "cred_avenue", "hindon_colending",
    "arthmate_i2ifunding", "kaleidofin_da1", "liquiloans", "janasha_colending",
}


# ── Superset auth ──────────────────────────────────────────────────────────
def _authenticate(session):
    login = session.post(
        f"{SUPERSET_URL}/api/v1/security/login",
        json={
            "username": session._superset_username,
            "password": session._superset_password,
            "provider": "db",
            "refresh": True,
        },
        timeout=60,
    )
    login.raise_for_status()
    token = login.json()["access_token"]
    csrf = session.get(
        f"{SUPERSET_URL}/api/v1/security/csrf_token/",
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    csrf.raise_for_status()
    session.headers.update({
        "Authorization": f"Bearer {token}",
        "X-CSRFToken": csrf.json()["result"],
        "Content-Type": "application/json",
        "Referer": f"{SUPERSET_URL}/sqllab/",
    })


def superset_session():
    username = os.environ.get("SUPERSET_USERNAME")
    password = os.environ.get("SUPERSET_PASSWORD")
    if not username or not password:
        raise RuntimeError("Set SUPERSET_USERNAME and SUPERSET_PASSWORD.")
    session = requests.Session()
    session._superset_username = username
    session._superset_password = password
    _authenticate(session)
    return session


# ── SQL runner ─────────────────────────────────────────────────────────────
def run_sql(session, db_id, sql, query_limit=PAGE_SIZE, retries=3):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            response = session.post(
                f"{SUPERSET_URL}/api/v1/sqllab/execute/",
                json={
                    "database_id": db_id,
                    "sql": sql,
                    "runAsync": False,
                    "queryLimit": query_limit,
                },
                timeout=300,
            )
            if response.status_code == 401:
                print("  Token expired (401), re-authenticating...")
                _authenticate(session)
                continue
            if response.status_code != 200:
                raise RuntimeError(response.text)
            payload = response.json()
            payload_str = json.dumps(payload)
            if "Token has expired" in payload_str:
                print("  Token expired in payload, re-authenticating...")
                _authenticate(session)
                continue
            if payload.get("error") or payload.get("errors"):
                raise RuntimeError(payload_str)
            return payload.get("data", [])
        except Exception as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(2 * attempt)
    raise RuntimeError(last_error)


# ── Helpers ────────────────────────────────────────────────────────────────
def normalize_row(row):
    normalized = {}
    for col in RISK_COLUMNS:
        value = row.get(col)
        if hasattr(value, "isoformat"):
            value = value.isoformat()
        normalized[col] = value
    return normalized


def modified_dpd(dpd, lender):
    dpd = dpd or 0
    if lender in FULL_ZERO_LENDERS:
        return 0
    if dpd >= 240:
        return 0
    if 180 <= dpd < 240 and lender in PARTIAL_LENDERS_180:
        return 0
    return dpd


# ── Data fetchers ──────────────────────────────────────────────────────────
def fetch_risk_dump(session):
    print(f"Fetching risk dump for {SNAPSHOT_DATE}...")
    rows = []
    offset = 0
    select_cols = ", ".join(RISK_COLUMNS)
    while True:
        sql = f"""
        SELECT {select_cols}
        FROM risk_portfolioriskdump
        WHERE date_created::date = DATE '{SNAPSHOT_DATE}'
        ORDER BY loanshare_id
        LIMIT {PAGE_SIZE} OFFSET {offset}
        """
        batch = run_sql(session, RISK_DB_ID, sql, query_limit=PAGE_SIZE)
        if not batch:
            break
        rows.extend(normalize_row(row) for row in batch)
        print(f"  fetched {len(rows):,} rows")
        if len(batch) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    print(f"Risk dump complete: {len(rows):,} rows")
    return rows


def fetch_interest_map(session, all_rows):
    active_ids = [
        int(r["loanshare_id"])
        for r in all_rows
        if r["user_status"] == "loan_disbursed"
        and r["writeoff_status"] is None
        and r["writeoff_action"] is None
        and r["loanshare_id"] is not None
    ]
    print(f"Fetching interest for {len(active_ids):,} active loanshares...")
    interest = {}
    for start in range(0, len(active_ids), INTEREST_BATCH):
        chunk = active_ids[start: start + INTEREST_BATCH]
        ids = ",".join(str(x) for x in chunk)
        sql = f"""
        SELECT loanshare_id, COALESCE(SUM(interest), 0) AS accrued_interest
        FROM loan_installment
        WHERE loanshare_id IN ({ids})
          AND due_date >= DATE '{DUE_START}'
          AND due_date <= DATE '{DUE_END}'
        GROUP BY loanshare_id
        """
        found = {
            int(row["loanshare_id"]): row["accrued_interest"] or 0
            for row in run_sql(session, KOSH_DB_ID, sql, query_limit=INTEREST_BATCH)
        }
        for ls_id in chunk:
            interest[ls_id] = found.get(ls_id, 0)
        print(f"  interest rows: {len(interest):,}/{len(active_ids):,}")
    print(f"Interest map complete: {len(interest):,} rows")
    return interest


# ── Removal logic ──────────────────────────────────────────────────────────
def identify_removed_rows(all_rows):
    nf_wo = [
        r for r in all_rows
        if r["lender"] == "narendra_finance"
        and (r["writeoff_status"] is not None or r["writeoff_action"] is not None)
        and (r["principal_outstanding"] or 0) > 0
    ]
    nf_wo_sorted = sorted(
        nf_wo,
        key=lambda r: (r["dpd"] or 0, r["principal_outstanding"] or 0),
        reverse=True,
    )
    target = 2e7
    running = 0
    remove_ids = set()
    removed_rows = []
    for row in nf_wo_sorted:
        if running >= target:
            break
        remove_ids.add(row["loanshare_id"])
        removed_rows.append(row)
        running += row["principal_outstanding"]
    print(f"Removing {len(remove_ids):,} loanshares | PO = Rs {running/1e7:.4f} cr")
    return remove_ids, removed_rows, running


# ── Build DataFrame (used by both Dataiku and GitHub Actions) ──────────────
def build_dataframe(all_rows, interest_map, remove_ids):
    print("Building dataframe...")
    records = []
    for row in all_rows:
        ls_id = row["loanshare_id"]
        if ls_id in remove_ids:
            continue
        user_status = row["user_status"]
        writeoff_status = row["writeoff_status"]
        writeoff_action = row["writeoff_action"]
        principal_outstanding = row["principal_outstanding"] or 0
        lender = row["lender"]
        is_active = (
            user_status == "loan_disbursed"
            and writeoff_status is None
            and writeoff_action is None
        )
        accrued_interest = (
            interest_map.get(int(ls_id), 0)
            if is_active and ls_id is not None
            else 0
        )
        final_pos = principal_outstanding + accrued_interest
        disbursal_date = row["disbursal_date"]
        db_month = str(disbursal_date)[:7] if disbursal_date else None

        records.append({
            "snapshot_date": SNAPSHOT_DATE,
            "id": row["id"],
            "loan_id": row["loan_id"],
            "loanshare_id": ls_id,
            "amount": row["amount"],
            "disbursal_date": disbursal_date,
            "db_month": db_month,
            "name": row["name"],
            "mobile": row["mobile"],
            "pan_number": row["pan_number"],
            "state": row["state"],
            "aadhaar_number": row["aadhaar_number"],
            "pincode": row["pincode"],
            "address": row["address"],
            "partner_loan_id": row["partner_loan_id"],
            "last_payment_date": row["last_payment_date"],
            "gender": row["gender"],
            "dob": row["dob"],
            "age": row["age"],
            "employment_type": row["employment_type"],
            "annual_salary": row["annual_salary"],
            "topup": row["topup"],
            "cibil_score": row["cibil_score"],
            "foir_score": row["foir_score"],
            "total_collection": row["total_collection"],
            "dpd": modified_dpd(row["dpd"], lender),
            "status": row["status"],
            "tenure": row["tenure"],
            "lender": lender,
            "purpose": row["purpose"],
            "channel": row["channel"],
            "user_status": user_status,
            "overdue_amount": row["overdue_amount"],
            "principal_outstanding": principal_outstanding,
            "tags_location": row["tags_location"],
            "relationship_manager": row["relationship_manager"],
            "greypool": row["greypool"],
            "emi": row["emi"],
            "date_created": row["date_created"],
            "last_updated": row["last_updated"],
            "writeoff_action": writeoff_action,
            "writeoff_status": writeoff_status,
            "fldg": row["fldg"],
            "last_success_payment_date": row["last_success_payment_date"],
            "overdue_principal": row["overdue_principal"],
            "accrued_interest": accrued_interest,
            "final_pos": final_pos,
            "created_at": pd.Timestamp.now(),
        })

    df = pd.DataFrame(records)
    print(f"  Rows kept  : {len(df):,}")
    print(f"  Total POS  : Rs {df['final_pos'].sum()/1e7:.2f} cr")
    return df


# ── run_pipeline — entry point for Dataiku ─────────────────────────────────
def run_pipeline():
    """Called by Dataiku recipe. Returns DataFrame + metadata."""
    session = superset_session()
    all_rows = fetch_risk_dump(session)
    print(f"Risk rows: {len(all_rows):,}")
    interest_map = fetch_interest_map(session, all_rows)
    print(f"Interest map rows: {len(interest_map):,}")
    remove_ids, removed_rows, po_removed = identify_removed_rows(all_rows)
    df = build_dataframe(all_rows, interest_map, remove_ids)
    return df, removed_rows, po_removed


# ── GitHub Actions only — Excel + Drive + Email ────────────────────────────
def style_header(sheet, fill):
    from openpyxl.styles import Alignment, Font
    font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = center


def build_workbook(all_rows, interest_map, remove_ids, removed_rows, running):
    from openpyxl.styles import PatternFill
    blue_fill = PatternFill("solid", fgColor="1F4E79")
    red_fill = PatternFill("solid", fgColor="C00000")

    workbook = openpyxl.Workbook()
    main = workbook.active
    main.title = f"{today.strftime('%b').lower()}_final"
    main.append(HEADERS)
    style_header(main, blue_fill)

    kept = 0
    total_pos = 0
    for row in all_rows:
        ls_id = row["loanshare_id"]
        if ls_id in remove_ids:
            continue
        user_status = row["user_status"]
        writeoff_status = row["writeoff_status"]
        writeoff_action = row["writeoff_action"]
        principal_outstanding = row["principal_outstanding"] or 0
        lender = row["lender"]
        is_active = (
            user_status == "loan_disbursed"
            and writeoff_status is None
            and writeoff_action is None
        )
        accrued_interest = (
            interest_map.get(int(ls_id), 0)
            if is_active and ls_id is not None
            else 0
        )
        final_pos = principal_outstanding + accrued_interest
        total_pos += final_pos
        kept += 1
        disbursal_date = row["disbursal_date"]
        db_month = str(disbursal_date)[:7] if disbursal_date else None
        main.append([
            row["id"], row["loan_id"], ls_id, row["amount"], disbursal_date, db_month,
            row["name"], row["mobile"], row["pan_number"], row["state"],
            row["aadhaar_number"], row["pincode"], row["address"], row["partner_loan_id"],
            row["last_payment_date"], row["gender"], row["dob"], row["age"],
            row["employment_type"], row["annual_salary"], row["topup"], row["cibil_score"],
            row["foir_score"], row["total_collection"], modified_dpd(row["dpd"], lender),
            row["status"], row["tenure"], lender, row["purpose"], row["channel"],
            user_status, row["overdue_amount"], principal_outstanding, row["tags_location"],
            row["relationship_manager"], row["greypool"], row["emi"], row["date_created"],
            row["last_updated"], writeoff_action, writeoff_status, row["fldg"],
            row["last_success_payment_date"], row["overdue_principal"],
            accrued_interest, final_pos,
        ])

    print(f"  Rows kept : {kept:,}")
    print(f"  Total POS : Rs {total_pos/1e7:.2f} cr")

    removed = workbook.create_sheet("removed_loanshares")
    removed.append(REM_HEADERS)
    style_header(removed, red_fill)
    for row in removed_rows:
        removed.append([
            row["loanshare_id"], row["loan_id"], row["amount"], row["disbursal_date"],
            row["name"], row["mobile"], row["pan_number"], row["lender"],
            row["user_status"], row["writeoff_status"], row["writeoff_action"],
            row["principal_outstanding"], row["dpd"], row["overdue_amount"],
            row["overdue_principal"], row["last_payment_date"],
        ])

    print(f"  Removed rows : {len(removed_rows):,}")
    BASE.mkdir(parents=True, exist_ok=True)
    workbook.save(OUT_PATH)
    print(f"Saved to {OUT_PATH}")
    return kept, total_pos


def split_excel():
    from openpyxl.styles import PatternFill
    print("Splitting Excel into parts...")
    blue_fill = PatternFill("solid", fgColor="1F4E79")
    file_size_mb = OUT_PATH.stat().st_size / 1e6
    part_size_target_mb = 6
    num_parts = max(3, int(file_size_mb / part_size_target_mb) + 1)
    print(f"  File size: {file_size_mb:.1f} MB → splitting into {num_parts} parts")
    source = openpyxl.load_workbook(OUT_PATH, read_only=True)
    main_sheet = source.active
    all_data = list(main_sheet.iter_rows(min_row=2, values_only=True))
    source.close()
    total_rows = len(all_data)
    chunk_size = (total_rows + num_parts - 1) // num_parts
    part_paths = []
    for part_num in range(1, num_parts + 1):
        start = (part_num - 1) * chunk_size
        end = min(start + chunk_size, total_rows)
        chunk = all_data[start:end]
        if not chunk:
            break
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"part_{part_num}"
        ws.append(HEADERS)
        style_header(ws, blue_fill)
        for row in chunk:
            ws.append(list(row))
        part_path = BASE / f"b2b_sheet_{SNAPSHOT_DATE}_part{part_num}of{num_parts}.xlsx"
        wb.save(part_path)
        size_mb = part_path.stat().st_size / 1e6
        print(f"  Part {part_num}: {len(chunk):,} rows | {size_mb:.1f} MB")
        part_paths.append(part_path)
    return part_paths


def upload_to_drive():
    print("Uploading to Google Drive...")
    service_account_info = json.loads(GDRIVE_SERVICE_ACCOUNT_JSON)
    credentials = service_account.Credentials.from_service_account_info(
        service_account_info,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    drive_service = build("drive", "v3", credentials=credentials)
    file_name = f"B2B Report — {today.strftime('%B %Y')}.xlsx"
    existing = drive_service.files().list(
        q=f"name='{file_name}' and '{GDRIVE_FOLDER_ID}' in parents and trashed=false",
        fields="files(id, name)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    for f in existing.get("files", []):
        try:
            drive_service.files().delete(
                fileId=f["id"], supportsAllDrives=True
            ).execute()
            print(f"  Deleted existing: {f['name']}")
        except Exception as e:
            print(f"  Could not delete {f['name']}: {e}")
    file_metadata = {"name": file_name, "parents": [GDRIVE_FOLDER_ID]}
    media = MediaFileUpload(
        str(OUT_PATH),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    uploaded = drive_service.files().create(
        body=file_metadata, media_body=media,
        fields="id, webViewLink", supportsAllDrives=True,
    ).execute()
    file_id = uploaded["id"]
    view_link = uploaded["webViewLink"]
    drive_service.permissions().create(
        fileId=file_id,
        body={"type": "anyone", "role": "reader"},
        supportsAllDrives=True,
    ).execute()
    print(f"  Uploaded: {file_name} → {view_link}")
    return view_link


def send_part_emails(kept, total_pos, removed_count, po_removed, drive_link, part_paths):
    print(f"Sending {len(part_paths)} part emails...")
    for i, part_path in enumerate(part_paths, 1):
        size_mb = part_path.stat().st_size / 1e6
        print(f"  Sending part {i}/{len(part_paths)} ({size_mb:.1f} MB)...")
        msg = MIMEMultipart()
        msg["From"] = GMAIL_ADDRESS
        msg["To"] = REPORT_RECIPIENT
        msg["Subject"] = (
            f"✅ B2B Monthly Report — {today.strftime('%B %Y')} "
            f"(Part {i} of {len(part_paths)})"
        )
        body = f"Hi Anurag,\n\nPlease find attached Part {i} of {len(part_paths)} of the B2B sheet for {today.strftime('%B %Y')}.\n\n"
        body += f"🔗 Full file on Google Drive: {drive_link}\n\n"
        if i == 1:
            body += f"""Summary:
  Snapshot Date     : {SNAPSHOT_DATE}
  Final Loanshares  : {kept:,}
  Removed           : {removed_count:,}
  PO Removed        : Rs {po_removed/1e7:.4f} cr
  Final Total POS   : Rs {total_pos/1e7:.2f} cr

"""
        body += f"This is an automated report generated on {today.isoformat()}."
        msg.attach(MIMEText(body, "plain"))
        with open(part_path, "rb") as f:
            attachment = MIMEApplication(f.read(), _subtype="xlsx")
            attachment.add_header(
                "Content-Disposition", "attachment", filename=part_path.name
            )
            msg.attach(attachment)
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_ADDRESS, REPORT_RECIPIENT, msg.as_string())
        print(f"  Part {i} sent.")
        time.sleep(2)
    print(f"All {len(part_paths)} emails sent.")


def send_drive_only_email(kept, total_pos, removed_count, po_removed, drive_link, error):
    print("Sending Drive-only fallback email...")
    msg = MIMEMultipart()
    msg["From"] = GMAIL_ADDRESS
    msg["To"] = REPORT_RECIPIENT
    msg["Subject"] = f"✅ B2B Monthly Report — {today.strftime('%B %Y')} (Drive link)"
    body = f"""Hi Anurag,

The cleaned B2B sheet for {today.strftime('%B %Y')} is ready.

📎 Download here: {drive_link}

Note: Email attachment failed. Reason: {error}

Summary:
  Snapshot Date     : {SNAPSHOT_DATE}
  Final Loanshares  : {kept:,}
  Removed           : {removed_count:,}
  PO Removed        : Rs {po_removed/1e7:.4f} cr
  Final Total POS   : Rs {total_pos/1e7:.2f} cr

This is an automated report generated on {today.isoformat()}.
"""
    msg.attach(MIMEText(body, "plain"))
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_ADDRESS, REPORT_RECIPIENT, msg.as_string())
    print("Fallback email sent.")


def cleanup():
    for f in [OUT_PATH]:
        if f.exists():
            f.unlink()
    for p in BASE.glob("*_part*of*.xlsx"):
        p.unlink()
    if BASE.exists():
        try:
            BASE.rmdir()
            BASE.parent.rmdir()
        except OSError:
            pass
    print("Cleaned up local files")


# ── Main — GitHub Actions entry point ─────────────────────────────────────
def main():
    BASE.mkdir(parents=True, exist_ok=True)
    session = superset_session()
    all_rows = fetch_risk_dump(session)
    print(f"Risk rows: {len(all_rows):,}")
    interest_map = fetch_interest_map(session, all_rows)
    print(f"Interest map rows: {len(interest_map):,}")
    remove_ids, removed_rows, po_removed = identify_removed_rows(all_rows)
    kept, total_pos = build_workbook(
        all_rows, interest_map, remove_ids, removed_rows, po_removed
    )
    drive_link = upload_to_drive()
    removed_count = len(removed_rows)
    try:
        part_paths = split_excel()
        send_part_emails(kept, total_pos, removed_count, po_removed, drive_link, part_paths)
    except Exception as e:
        print(f"Split/attach failed: {e}")
        send_drive_only_email(kept, total_pos, removed_count, po_removed, drive_link, str(e))
    cleanup()
    print()
    print("=== SUMMARY ===")
    print(f"Snapshot date       : {SNAPSHOT_DATE}")
    print(f"Original loanshares : {len(all_rows):,}")
    print(f"Removed             : {len(removed_rows):,}")
    print(f"Final loanshares    : {kept:,}")
    print(f"PO removed          : Rs {po_removed/1e7:.4f} cr")
    print(f"Final Total POS     : Rs {total_pos/1e7:.2f} cr")


if __name__ == "__main__":
    main()
